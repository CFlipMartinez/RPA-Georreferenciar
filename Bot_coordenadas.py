import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys #para enviar teclas especiales
from selenium.webdriver.support.ui import WebDriverWait #para manejar tiempos de espera de carga de la web
from selenium.webdriver.support import expected_conditions as EC #se usa para esperar condiciones
import pandas as pd
import time
from openpyxl import load_workbook
import os
import folium

def georeferenciar():
    """LECTURA DE DATOS"""
    # rutaExcel = os.path.expanduser('~').replace('\\','/')+'/Desktop/Trabajos U/Ingenieria desarrollo de sotware/semestre I/PIA/Georeferenciacion/Bot_Coordenadas/Formato/Formato.xlsx'
    """CUADRO DE DIÁLOGO PARA SELECCIONAR EL ARCHIVO EXCEL"""
    rutaExcel = filedialog.askopenfilename(title="Selecciona el archivo Excel", filetypes=[("Archivos de Excel", "*.xlsx")])

    # Verifica si se seleccionó un archivo
    if not rutaExcel:
        alert = tk.Tk()
        alert.withdraw()  # Ocultar la ventana principal
        messagebox.showinfo("Atención", "No se seleccionó ningún archivo. Cerrando programa")
        alert.destroy()  # Cerrar la ventana emergente
        exit()

    #contar cantidad de registros
    filesheet = rutaExcel
    wb = load_workbook(filesheet)
    sheet = wb.active
    rowExcel=1 
    rowExcelString=str(rowExcel)

    num_filas = sheet.max_row
    num_columnas = sheet.max_column
    ref = 'A'

    cont = 0
    for fila in range(1, num_filas + 1):
        celda = f'{ref}{fila}'
        if sheet[celda].value is not None:
            cont += 1
    cont=cont+1
    print("Numero de filas en el Excel: ",cont-1)

    #Definicon inicial de variables cambiantes

    FilasExcel = 2

    # Crear un mapa de Folium
    mapa = None

    """INICIAR GOOGLE CHROME"""
    url =  'https://www.google.com/maps'
    driver = webdriver.Chrome() #Abre Google chrome
    driver.maximize_window() #maximiza la pestaña

    """PROCESO EXTRACCION DE COORDENADAS"""

    for i in range(2,cont):
        #Cargar datos excel
        filesheet=rutaExcel
        wb = load_workbook(filesheet)
        sheet = wb.active
            
        cell_A = sheet[f'A{i}']
        cell_B = sheet[f'B{i}']
        cell_C = sheet[f'C{i}']
        cell_D = sheet[f'D{i}']
        cell_E = sheet[f'E{i}']
        cell_F = sheet[f'F{i}']

        Ciudad = cell_B
        nombreLugar = cell_C
        Direccion = cell_D
        Latitud = cell_E
        Longitud = cell_F

        Ciudad = str(Ciudad.value)
        Ciudad = Ciudad.strip()
        nombreLugar = str(nombreLugar.value)
        nombreLugar = nombreLugar.strip()
        Direccion = str(Direccion.value)
        Direccion = Direccion.strip()
        Latitud = str(Latitud.value)
        Latitud = Latitud.strip()
        Longitud = str(Longitud.value)
        Longitud = Longitud.strip()

        driver.get(url) #ingresa a url
        """Busqueda de direcciones"""
        adress = str(Direccion + ', ' + nombreLugar + ', ' + Ciudad) #crear la llave de busqueda para cda dirección
        wait = WebDriverWait(driver, 120, 5)
        search_box = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="searchboxinput"]')))
        search_box = driver.find_element(By.XPATH, '//*[@id="searchboxinput"]') #localiza el cuadro de busqueda de google Maps
        #verificar que el cuadro de busqueda esté vacio y luego escrie la dirección
        if search_box is None:
            search_box.send_keys(adress)
        else:
            search_box.clear()
            search_box.send_keys(adress)
        print('Dirección: ',adress)
        wait = WebDriverWait(driver, 120, 5)
        search = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchbox-searchbutton"]')))
        search.click() #click en el botón de busqueda
        time.sleep(5)

        """Extracción de cooredenadas"""
        urlCoordenadas = driver.current_url #obtiene la url actual
        lat_long = urlCoordenadas.split('@')[1].split(',')[0:2] #divide la url para obtener las cooredenadas
        lat, long = map(str, lat_long) #guarda en cada variable latitud y longitud en tipo de dato str
        Latitud = lat #agrega el valor obtenido de latitud en la columna latitud del exel 
        Longitud = long #agrega el valor obtenido de longitud en la columna longitud del exel
        filesheet = rutaExcel
        wb = load_workbook(filesheet)
        sheet = wb.active
        rowExcelString=str(FilasExcel)
        sheet['E'+rowExcelString] = Latitud
        sheet['F'+rowExcelString] = Longitud
        time.sleep(2)
        wb.save(filesheet)
        wb.close
        time.sleep(2)
        FilasExcel = int(FilasExcel+1)
        time.sleep(2)

        if mapa is None:
            # Definir las coordenadas iniciales y el nivel de zoom
            mapa = folium.Map(location=[lat, long], zoom_start=15)

        # Agregar marcador al mapa de Folium
        folium.Marker([lat, long], popup=f'Dirección: {adress}').add_to(mapa)

    # Guardar el mapa en un archivo HTML
    if mapa is not None:
        mapa.save(os.path.expanduser('~').replace('\\', '/') + '/Documents/Universidad/PIA/Bot_Coordenadas/Mapa/mapa_coordenadas.html')


    driver.quit() #cierra el navegador una vez termina el ciclo
    # Mostrar alerta de finalización
    alert = tk.Tk()
    alert.withdraw()  # Ocultar la ventana principal
    messagebox.showinfo("Finalización", "El proceso ha finalizado, direcciones georreferenciadas.")
    alert.destroy()  # Cerrar la ventana emergente


# Crear la interfaz gráfica
root = tk.Tk()
root.title("GeoPlanner")
root.geometry("250x90") #Ajustar tamaño
root.resizable(0,0)

# Crear el botón "Georeferenciar"
boton_georreferenciar = tk.Button(root, text="Georeferenciar", command=georeferenciar)
boton_georreferenciar.pack(pady=20)

# Iniciar la interfaz gráfica
root.mainloop()








